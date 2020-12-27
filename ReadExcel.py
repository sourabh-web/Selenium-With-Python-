import openpyxl

path="C:\\Downloadfiles\\Book2.xlsx"

workbook=openpyxl.load_workbook(path)

sheet=workbook.active      #workbook.get_sheet_by_name("sheet1")

row=sheet.max_row
colmn=sheet.max_column

print(row)
print(colmn)

for r in range(1,row+1):
    for c in range(1,colmn+1):
        print(sheet.cell(row=r,column=c).value,end="  ")

    print()
